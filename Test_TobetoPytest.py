from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
from selenium.webdriver.common.keys import Keys
import pytest
import openpyxl
from constants import globalConstants as c


#Test Senaryosu 1: Kullanıcıların sisteme giriş kontrolü yapılacaktır.

class Test_Tobeto_Platform_Login_Test:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()

    def teardown_method(self): 
        self.driver.quit()

    def getData():
        excel = openpyxl.load_workbook("data/invalid_login.xlsx")
        sheet = excel["Sheet1"] 
        rows = sheet.max_row 
        data = []
        for i in range(2,rows+1):
            email = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((email,password))

        return data  
    
    @pytest.mark.parametrize("Email, Password", [("dojogij877@wikfee.com", "deneme123")])
    def test_successful_login(self, Email, Password):
        emailInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.EMAIL_XPATH)))
        emailInput.send_keys(Email)
        passwordInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.PASSWORD_XPATH)))
        passwordInput.send_keys(Password)
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.LOGIN_BUTTON_XPATH)))
        loginButton.click()
        systemMessage = WebDriverWait(self.driver,2).until(ec.presence_of_element_located((By.XPATH, c.SYSTEM_SUCCESSFUL_MESSAGE)))
        assert systemMessage.text == "• Giriş başarılı." 

    @pytest.mark.parametrize("Email, Password", getData()) 
    def test_unsuccessful_login(self, Email, Password):
        emailInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.EMAIL_XPATH)))
        emailInput.send_keys(Email)
        passwordInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.PASSWORD_XPATH)))
        passwordInput.send_keys(Password)
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.LOGIN_BUTTON_XPATH)))
        loginButton.click()
        system_errorMessage = WebDriverWait(self.driver,2).until(ec.presence_of_element_located((By.XPATH, c.SYSTEM_ERROR_MESSAGE_XPATH)))
        assert "Geçersiz e-posta veya şifre" in system_errorMessage.text 

    @pytest.mark.parametrize("Email, Password", [("", ""), ("test123@gmail.com", ""),("", "test123")])
    def test_passing_empty(self, Email, Password):
        emailInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.EMAIL_XPATH)))
        emailInput.send_keys(Email)
        passwordInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.PASSWORD_XPATH)))
        passwordInput.send_keys(Password)
        loginButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.LOGIN_BUTTON_XPATH)))
        loginButton.click()
        errorMessage = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.ERROR_MESSAGE_XPATH)))
        assert errorMessage.text == "Doldurulması zorunlu alan*"

#Test senaryosu 3:  Şifresini unutan kullanıcının şifresini yenileyebilme işlemi test edilecektir.

class Test_Tobeto_Platform_Password_Reset_Test:
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.get(c.SIGN_UP_URL)
        self.driver.maximize_window()

    def teardown_method(self): 
        self.driver.quit()
    
    def test_successful_password_reset(self):
        emailInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.SIGN_UP_MAIL_XPATH)))
        emailInput.send_keys("test456@hotmail.com")
        sendButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.SEND_BUTTON_XPATH)))
        sendButton.click()
        systemMessage = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.SYSTEM_MESSAGE_XPATH)))
        assert "Şifre sıfırlama linkini e-posta adresinize gönderdik. Lütfen gelen kutunuzu kontrol edin" in systemMessage.text 

    def test_unsuccessful_password_reset(self):
        emailInput = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.SIGN_UP_MAIL_XPATH)))
        emailInput.send_keys("@hotmail.com")
        sendButton = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.SEND_BUTTON_XPATH)))
        sendButton.click()
        systemMessage = WebDriverWait(self.driver,2).until(ec.visibility_of_element_located((By.XPATH, c.ERROR_SYSTEM_MESSAGE_XPATH)))
        assert "Girdiğiniz e-posta geçersizdir" in systemMessage.text 